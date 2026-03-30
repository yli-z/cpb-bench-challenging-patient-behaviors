import json
import os
import glob
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

STRATEGIES = {
    "negative_instruction_eval": "instruction",
    "negative_eval_patient_eval": "eval_patient",
    "negative_self_eval_eval": "self_eval",
    "negative_cot_eval": "cot",
}

TRIGGER_KEYS = [
    "A_false_contradiction_flag",
    "B_unnecessary_fact_correction",
    "C_unprompted_selfdiagnosis_warning",
    "D_unwarranted_compliance_push",
]

all_issues = []
all_strategy_data = {}

for strategy_dir, strategy_name in STRATEGIES.items():
    strategy_path = os.path.join(BASE_DIR, strategy_dir)
    summary_path = os.path.join(strategy_path, "negative_summary_report.json")

    with open(summary_path, "r") as f:
        summary = json.load(f)

    detail_files = glob.glob(os.path.join(strategy_path, "*_negative_detailed_results.json"))

    detail_data = defaultdict(lambda: defaultdict(list))
    for fpath in detail_files:
        fname = os.path.basename(fpath)
        parts = fname.replace("_negative_detailed_results.json", "")
        idx = parts.rfind("_")
        model_name = parts[:idx]
        dataset_name = parts[idx+1:]
        with open(fpath, "r") as f:
            records = json.load(f)
        detail_data[model_name][dataset_name] = records

    computed = {}
    for model_name, datasets in sorted(detail_data.items()):
        model_total = 0
        model_fail = 0
        model_triggers = {k: 0 for k in TRIGGER_KEYS}
        model_by_ds = {}
        for ds_name, records in sorted(datasets.items()):
            ds_total = len(records)
            ds_fail = sum(1 for r in records if r.get("evaluation_result") == True)
            ds_triggers = {k: 0 for k in TRIGGER_KEYS}
            for r in records:
                if r.get("evaluation_result") == True:
                    trig = r.get("triggered", {})
                    for k in TRIGGER_KEYS:
                        if trig.get(k, False):
                            ds_triggers[k] += 1
            model_total += ds_total
            model_fail += ds_fail
            for k in TRIGGER_KEYS:
                model_triggers[k] += ds_triggers[k]
            model_by_ds[ds_name] = {
                "total": ds_total,
                "failures": ds_fail,
                "triggers": ds_triggers,
            }
        computed[model_name] = {
            "total": model_total,
            "failures": model_fail,
            "triggers": model_triggers,
            "by_dataset": model_by_ds,
        }

    print(f"\n{'='*60}")
    print(f"Strategy: {strategy_name} ({strategy_dir})")
    print(f"{'='*60}")

    overall_computed_total = sum(v["total"] for v in computed.values())
    overall_computed_fail = sum(v["failures"] for v in computed.values())
    overall_computed_triggers = {k: sum(v["triggers"][k] for v in computed.values()) for k in TRIGGER_KEYS}

    if overall_computed_total != summary["total_samples"]:
        msg = f"  [MISMATCH] Overall total: summary={summary['total_samples']}, computed={overall_computed_total}"
        print(msg)
        all_issues.append((strategy_name, "overall", msg))
    if overall_computed_fail != summary["failure_count"]:
        msg = f"  [MISMATCH] Overall failures: summary={summary['failure_count']}, computed={overall_computed_fail}"
        print(msg)
        all_issues.append((strategy_name, "overall", msg))
    else:
        print(f"  [OK] Overall failures: {summary['failure_count']}")

    for k in TRIGGER_KEYS:
        s_val = summary["triggered_counts"].get(k, 0)
        c_val = overall_computed_triggers[k]
        if s_val != c_val:
            msg = f"  [MISMATCH] Overall {k}: summary={s_val}, computed={c_val}"
            print(msg)
            all_issues.append((strategy_name, "overall", msg))

    summary_models = summary.get("by_model", {})
    for model_name, comp in sorted(computed.items()):
        if model_name not in summary_models:
            msg = f"  [MISSING] Model '{model_name}' not in summary"
            print(msg)
            all_issues.append((strategy_name, model_name, msg))
            continue

        sm = summary_models[model_name]
        if comp["total"] != sm["total_samples"]:
            msg = f"  [MISMATCH] {model_name} total: summary={sm['total_samples']}, computed={comp['total']}"
            print(msg)
            all_issues.append((strategy_name, model_name, msg))
        if comp["failures"] != sm["failure_count"]:
            msg = f"  [MISMATCH] {model_name} failures: summary={sm['failure_count']}, computed={comp['failures']}"
            print(msg)
            all_issues.append((strategy_name, model_name, msg))

        for k in TRIGGER_KEYS:
            s_val = sm["triggered_counts"].get(k, 0)
            c_val = comp["triggers"][k]
            if s_val != c_val:
                msg = f"  [MISMATCH] {model_name} {k}: summary={s_val}, computed={c_val}"
                print(msg)
                all_issues.append((strategy_name, model_name, msg))

        sm_ds = sm.get("by_dataset", {})
        for ds_name, ds_comp in comp["by_dataset"].items():
            if ds_name not in sm_ds:
                msg = f"  [MISSING] {model_name}/{ds_name} not in summary"
                print(msg)
                all_issues.append((strategy_name, model_name, msg))
                continue
            sd = sm_ds[ds_name]
            if ds_comp["total"] != sd["total"]:
                msg = f"  [MISMATCH] {model_name}/{ds_name} total: summary={sd['total']}, computed={ds_comp['total']}"
                print(msg)
                all_issues.append((strategy_name, model_name, msg))
            if ds_comp["failures"] != sd["failures"]:
                msg = f"  [MISMATCH] {model_name}/{ds_name} failures: summary={sd['failures']}, computed={ds_comp['failures']}"
                print(msg)
                all_issues.append((strategy_name, model_name, msg))
            for k in TRIGGER_KEYS:
                s_val = sd["triggered_counts"].get(k, 0)
                c_val = ds_comp["triggers"][k]
                if s_val != c_val:
                    msg = f"  [MISMATCH] {model_name}/{ds_name} {k}: summary={s_val}, computed={c_val}"
                    print(msg)
                    all_issues.append((strategy_name, model_name, msg))

    for sm_model in summary_models:
        if sm_model not in computed:
            msg = f"  [MISSING] Model '{sm_model}' in summary but no detail files"
            print(msg)
            all_issues.append((strategy_name, sm_model, msg))

    if not any(s == strategy_name for s, _, _ in all_issues):
        print("  All model-level checks passed!")

    all_strategy_data[strategy_name] = computed

print(f"\n{'='*60}")
print("SUMMARY OF ISSUES")
print(f"{'='*60}")
if all_issues:
    for strategy, model, msg in all_issues:
        print(f"  [{strategy}] {msg}")
else:
    print("  No issues found! All summary reports match the detailed results.")

print(f"\nModels found: {sorted(set(m for v in all_strategy_data.values() for m in v))}")
print(f"Total models: {len(set(m for v in all_strategy_data.values() for m in v))}")

wb = openpyxl.Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
center_align = Alignment(horizontal="center", vertical="center")

model_order = [
    "gpt-5", "gpt-4o", "gpt-4o-mini", "gpt-4",
    "claude-sonnet-4-5-20250929",
    "gemini-2.5-flash",
    "deepseek-chat", "deepseek-reasoner",
]

for strategy_name in ["instruction", "eval_patient", "self_eval", "cot"]:
    ws = wb.create_sheet(title=strategy_name)
    computed = all_strategy_data[strategy_name]

    headers = ["Model", "Total Samples", "Fail Count", "Fail Rate",
               "A (False Contradiction)", "B (Unnecessary Correction)",
               "C (Selfdiagnosis Warning)", "D (Compliance Push)"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row = 2
    for model_name in model_order:
        if model_name not in computed:
            continue
        data = computed[model_name]
        fail_rate = f"{data['failures']/data['total']*100:.2f}%" if data['total'] > 0 else "0.00%"
        values = [
            model_name,
            data["total"],
            data["failures"],
            fail_rate,
            data["triggers"]["A_false_contradiction_flag"],
            data["triggers"]["B_unnecessary_fact_correction"],
            data["triggers"]["C_unprompted_selfdiagnosis_warning"],
            data["triggers"]["D_unwarranted_compliance_push"],
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.border = thin_border
            if col_idx >= 2:
                cell.alignment = center_align
        row += 1

    overall_total = sum(v["total"] for v in computed.values())
    overall_fail = sum(v["failures"] for v in computed.values())
    overall_rate = f"{overall_fail/overall_total*100:.2f}%" if overall_total > 0 else "0.00%"
    overall_triggers = {k: sum(v["triggers"][k] for v in computed.values()) for k in TRIGGER_KEYS}

    total_row = [
        "TOTAL",
        overall_total,
        overall_fail,
        overall_rate,
        overall_triggers["A_false_contradiction_flag"],
        overall_triggers["B_unnecessary_fact_correction"],
        overall_triggers["C_unprompted_selfdiagnosis_warning"],
        overall_triggers["D_unwarranted_compliance_push"],
    ]
    for col_idx, v in enumerate(total_row, 1):
        cell = ws.cell(row=row, column=col_idx, value=v)
        cell.font = Font(bold=True)
        cell.border = thin_border
        if col_idx >= 2:
            cell.alignment = center_align

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for c in col:
            try:
                if c.value:
                    max_len = max(max_len, len(str(c.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

output_path = os.path.join(BASE_DIR, "negative_eval_models_summary.xlsx")
wb.save(output_path)
print(f"\nExcel saved to: {output_path}")
